import os
import re
import traceback
from datetime import datetime
import numpy as np
import cv2
import pytesseract
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from pdf2image import convert_from_path
import pandas as pd
import io
from thefuzz import fuzz
from sqlalchemy import or_
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import tempfile
from flask import send_file
from collections import OrderedDict
from datetime import datetime
import hashlib
from io import BytesIO

# ==============================================================================
# KONFIGURASI APLIKASI
# ==============================================================================
app = Flask(__name__)
CORS(app, origins=["http://localhost:3000"])

app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:password@localhost/proyek_pajak'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
# Path ke folder bin dari poppler di Windows
POPPLER_PATH = r'C:\poppler\poppler-24.08.0\Library\bin'

class PpnMasukan(db.Model):
    __tablename__ = 'ppn_masukan'
    id = db.Column(db.Integer, primary_key=True)
    bulan = db.Column(db.String(20), nullable=False)
    tanggal = db.Column(db.Date, nullable=False)
    keterangan = db.Column(db.Text, nullable=True)
    npwp_lawan_transaksi = db.Column(db.String(100), nullable=False)
    nama_lawan_transaksi = db.Column(db.String(255), nullable=False)
    no_faktur = db.Column(db.String(100), unique=True, nullable=False)
    dpp = db.Column(db.Numeric(15, 2), nullable=False)
    ppn = db.Column(db.Numeric(15, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())

# Model untuk tabel PPN Keluaran
class PpnKeluaran(db.Model):
    __tablename__ = 'ppn_keluaran'
    id = db.Column(db.Integer, primary_key=True)
    bulan = db.Column(db.String(20), nullable=False)
    tanggal = db.Column(db.Date, nullable=False)
    keterangan = db.Column(db.Text, nullable=True)
    npwp_lawan_transaksi = db.Column(db.String(100), nullable=False)
    nama_lawan_transaksi = db.Column(db.String(255), nullable=False)
    no_faktur = db.Column(db.String(100), unique=True, nullable=False)
    dpp = db.Column(db.Numeric(15, 2), nullable=False)
    ppn = db.Column(db.Numeric(15, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

# ==============================================================================
# FUNGSI HELPER
# ==============================================================================
def clean_number(text):
    """Membersihkan string angka dari format Rupiah ke format standar float."""
    if not text: return 0.0
    cleaned_text = re.sub(r'[^\d,.-]', '', text).strip()
    try:
        # Jika format Indonesia: 1.234.567,89 → 1234567.89
        if ',' in cleaned_text and '.' in cleaned_text:
            cleaned_text = cleaned_text.replace('.', '').replace(',', '.')
        # Jika format: 1.000.000 → 1000000
        elif '.' in cleaned_text and ',' not in cleaned_text:
            cleaned_text = cleaned_text.replace('.', '')
        # Jika format: 1,000.00 (gaya US), tidak diubah

        return float(cleaned_text)
    except ValueError:
        print(f"[SKIP] clean_number gagal mengubah '{text}' → None")
        return 0.0

def clean_string(text):
    """
    Membersihkan string nama perusahaan untuk perbandingan yang andal,
    dengan penanganan noise OCR yang lebih baik.
    """
    if not text:
        return ""

    # 1. Jika ada ':', ambil teks setelahnya untuk membuang label seperti "Nama :"
    if ':' in text:
        text = text.split(':', 1)[1]

    text = text.upper()

    # 2. Hapus semua simbol selain huruf dan spasi. Ini lebih agresif dari \w
    #    karena \w juga mencakup angka dan _, yang jarang ada di awal/tengah nama PT.
    text = re.sub(r'[^A-Z\s]', '', text)

    # 3. Hapus sebutan umum perusahaan
    text = re.sub(r'\b(PT|CV|TBK|PERSERO|PERUM|UD)\b', '', text)
    
    # 4. Pisahkan menjadi kata-kata dan buang noise
    words = text.split()
    
    # Simpan hanya kata-kata yang panjangnya lebih dari 2 huruf,
    # ini efektif menghilangkan noise seperti 'AN', 'AA', 'MN'
    cleaned_words = [word for word in words if len(word) > 2]

    # 5. Gabungkan kembali kata-kata yang sudah bersih
    return " ".join(cleaned_words).strip()

def extract_faktur_tanggal(raw_text):
    import re
    from datetime import datetime

    no_faktur = None
    tanggal_obj = None

    print("[🧾 DEBUG OCR-Easy Input] ------------")
    print(raw_text)
    print("--------------------------------------")

    # Normalisasi koma → titik, dan hilangkan spasi tidak perlu
    normalized_text = raw_text.replace(",", ".").replace("“", '"')

    # Ekstrak semua kemungkinan kandidat faktur dengan regex toleran
    kandidat_faktur = re.findall(r"\d{3}[.\s]?\d{3}[-.\s]?\d{2}[.\s]?\d{8}", normalized_text)
    kandidat_faktur = [re.sub(r"\s+", "", k) for k in kandidat_faktur]

    if kandidat_faktur:
        no_faktur = max(kandidat_faktur, key=len)
        print("[✅ DEBUG] Nomor faktur ditemukan:", no_faktur)
    else:
        print("[❌ DEBUG] Nomor faktur tidak ditemukan")

    # Ekstraksi tanggal
    tanggal_match = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", raw_text)
    if tanggal_match:
        hari, bulan, tahun = tanggal_match.groups()
        bulan_map = {
            "januari": "January", "februari": "February", "maret": "March", "april": "April",
            "mei": "May", "juni": "June", "juli": "July", "agustus": "August", "september": "September",
            "oktober": "October", "november": "November", "desember": "December"
        }
        bulan_inggris = bulan_map.get(bulan.lower())
        if bulan_inggris:
            try:
                tanggal_obj = datetime.strptime(f"{hari} {bulan_inggris} {tahun}", "%d %B %Y")
                print("[DEBUG] Tanggal ditemukan:", tanggal_obj)
            except Exception as e:
                print("[❌ ERROR] Gagal parsing tanggal:", e)
        else:
            print("[❌ DEBUG] Nama bulan tidak dikenali:", bulan)

    return no_faktur, tanggal_obj

def format_currency(value, with_symbol=True):
    try:
        value = float(value)
        formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"Rp {formatted}" if with_symbol else formatted
    except:
        return "Rp 0,00" if with_symbol else "0,00"

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_image_file(filename):
    return filename.lower().endswith(('.jpg', '.jpeg', '.png'))

def is_valid_image(path):
    try:
        with Image.open(path) as img:
            img.verify()
        return True
    except Exception:
        return False

def extract_keterangan(raw_text):
    try:
        start_match = re.search(r"Nama\s+Barang\s+Kena\s+Pajak.*?", raw_text, re.IGNORECASE)
        end_match = re.search(r"Dasar\s+Pengenaan\s+Pajak", raw_text, re.IGNORECASE)

        if not start_match or not end_match:
            return "Tidak ditemukan"

        block = raw_text[start_match.end():end_match.start()]
        lines = [line.strip() for line in block.splitlines() if line.strip()]

        cleaned_lines = []
        seen_lines = set()

        # Kata-kata noise umum dari OCR yang tidak relevan
        noise_words = {'oh', 'ka', 'bah', 'iai', 'aa', 'tr', 'id', 'na', 'in', '5', '2', '3', '4', 'es', 'po', 'sz'}
        typo_map = {
            "DECA R": "DECANTER", 
            "DESAND YCLON": "DESANDING CYCLONE",
            "PESIFIKA -SUA": "SPESIFIKASI SESUAI",
            "MATERI Tera": "MATERIAL", 
            "MATER INSTALASI": "MATERIAL INSTALASI",
            "ikurangi": "Dikurangi",  # typo umum
        }

        for line in lines:
            if line in seen_lines:
                continue
            seen_lines.add(line)

            # Normalisasi dasar
            line = re.sub(r'[^\w\s.,:;/\-()Rp]', '', line).strip()

            # Koreksi typo
            for typo, correct in typo_map.items():
                if typo in line:
                    line = line.replace(typo, correct)

            # Hilangkan token noise (kata satuan pendek tak relevan)
            tokens = [tok for tok in line.split() if tok.lower() not in noise_words]
            if not tokens:
                continue

            # Gabung ulang dan tambahkan separator
            cleaned_line = ' '.join(tokens).strip()
            if cleaned_line:
                cleaned_lines.append(cleaned_line)

        return ' || '.join(cleaned_lines) if cleaned_lines else "Tidak ditemukan"

    except Exception as e:
        print(f"[ERROR extract_keterangan] {e}")
        return "Tidak ditemukan"
    
def extract_jenis_pajak(raw_text, pt_utama):
    print("[DEBUG] Mulai extract_jenis_pajak...")
    pt_clean = clean_string(pt_utama)
    parts = re.split(r"Pembeli\s+Barang\s+Kena\s+Pajak", raw_text, flags=re.IGNORECASE)
    if len(parts) < 2:
        print("[DEBUG] ❌ Bagian 'Pembeli Barang Kena Pajak' tidak ditemukan, fallback ke full text.")
        for line in raw_text.splitlines():
            if fuzz.ratio(clean_string(line), pt_clean) > 70:
                print(f"[DEBUG] ✅ Ditemukan nama PT utama di full text → PPN MASUKAN")
                return 'PPN_MASUKAN', "", raw_text
        print("[DEBUG] ❌ Nama PT utama tidak ditemukan di fallback.")
        return None, None, None

    blok_penjual, blok_pembeli = parts
    print("[DEBUG] Bagian pembeli dan penjual ditemukan.")
    print("------ Blok Pembeli ------")
    print(blok_pembeli)
    print("------ Blok Penjual ------")
    print(blok_penjual)

    for line in blok_pembeli.splitlines():
        if fuzz.ratio(clean_string(line), pt_clean) > 70:
            print(f"[DEBUG] ✅ Nama PT cocok di blok pembeli: {line}")
            return 'PPN_MASUKAN', blok_penjual, blok_pembeli

    for line in blok_penjual.splitlines():
        if fuzz.ratio(clean_string(line), pt_clean) > 70:
            print(f"[DEBUG] ✅ Nama PT cocok di blok penjual: {line}")
            return 'PPN_KELUARAN', blok_pembeli, blok_penjual

    print("[DEBUG] ❌ Nama PT tidak cocok di kedua blok.")
    return None, None, None

def extract_npwp_nama_rekanan(blok_text):
    """
    Menangkap nama dan NPWP rekanan dari blok teks pembeli/penjual
    dengan logika pembersihan noise yang lebih baik.
    """
    nama = "Tidak Ditemukan"
    npwp = "Tidak Ditemukan"
    lines = blok_text.splitlines()

    for line in lines:
        line_clean = line.lower().strip()

        # Ekstraksi Nama dengan pembersihan noise
        if 'nama' in line_clean:
            # Ambil semua teks setelah ':'
            parts = line.split(':', 1)
            if len(parts) == 2:
                nama_candidate = parts[1].strip()
                
                # Hapus noise umum dan karakter non-alfanumerik di awal
                nama_candidate = re.sub(r'^[^\w]+', '', nama_candidate)
                
                # Pisahkan kata-kata, bersihkan, dan gabungkan kembali.
                # Ini efektif menghilangkan noise seperti 'AN an aa Mn'
                words = nama_candidate.split()
                cleaned_words = []
                for word in words:
                    # Hanya pertahankan kata yang mengandung setidaknya satu huruf
                    # dan memiliki panjang yang wajar.
                    if any(c.isalpha() for c in word) and len(word) > 1:
                         # Hapus simbol yang tidak relevan dari kata
                        cleaned_word = re.sub(r'[^A-Za-z0-9.&]', '', word)
                        if cleaned_word:
                            cleaned_words.append(cleaned_word)
                
                # Ambil beberapa kata pertama yang paling mungkin adalah nama PT
                # (misalnya, maksimal 5 kata pertama)
                nama = ' '.join(cleaned_words[:5])

        # Ekstraksi NPWP (logika Anda sudah cukup baik)
        if 'npwp' in line_clean:
            digits = re.sub(r'\D', '', line)
            if len(digits) >= 15:
                npwp15 = digits[:15]
                npwp = f"{npwp15[:2]}.{npwp15[2:5]}.{npwp15[5:8]}.{npwp15[8]}.{npwp15[9:12]}.{npwp15[12:15]}"

    # Pastikan nama yang dikembalikan tidak kosong
    return nama.strip() if nama.strip() else "Tidak Ditemukan", npwp.strip()

def extract_dpp(raw_text):
    try:
        dpp = 0.0
        dpp_line = ""

        lines = raw_text.splitlines()
        for line in lines:
            if 'dasar pengenaan pajak' in line.lower():
                numbers = re.findall(r'[\d.,]+', line)
                if numbers:
                    last_number = numbers[-1]
                    dpp = clean_number(last_number)
                    dpp_line = line
                    print(f"[✅ DPP dari baris] {dpp:,.2f} ← {line}")
                    break

        # Ambil semua angka besar
        all_numbers = re.findall(r"[\d.]{1,3}(?:[.,]\d{3}){2,}", raw_text)
        candidates = [clean_number(n) for n in all_numbers if clean_number(n) > 10_000_000]

        if dpp > 0:
            # Jika ada kandidat yang terlalu jauh dari DPP, abaikan override
            if candidates:
                max_val = max(candidates)
                if max_val > dpp * 5:
                    print(f"[❌ Abaikan fallback DPP] {max_val:,.2f} terlalu jauh dari {dpp:,.2f}")
                    return dpp, format_currency(dpp), None, None

            return dpp, format_currency(dpp), None, None

        # Kalau tidak ada "Dasar Pengenaan Pajak", fallback ke kandidat
        if candidates:
            fallback = max(candidates)
            print(f"[⚠️ Fallback DPP] {fallback:,.2f}")
            return fallback, format_currency(fallback), None, None

        return 0.0, format_currency(0.0), None, None

    except Exception as e:
        print(f"[ERROR extract_dpp] {e}")
        return 0.0, format_currency(0.0), None, None
    
def extract_ppn(raw_text, dpp, override_ppn=None):
    try:
        if override_ppn:
            print(f"[✅ Override PPN] {override_ppn:,.2f}")
            return override_ppn, format_currency(override_ppn)

        ppn = 0.0
        harga_jual = None
        angka_terdeteksi = []

        lines = raw_text.splitlines()
        relevant_lines = []
        found_trigger = False
        for line in lines:
            if not found_trigger and re.search(r'Barang|Jasa|Harga Jual', line, re.IGNORECASE):
                found_trigger = True
            if found_trigger:
                relevant_lines.append(line)

        total_ppn_match = re.search(r'Total\s*PPN\s*[:\-]?\s*([\d.,]+)', raw_text, re.IGNORECASE)
        if total_ppn_match:
            ppn_val = clean_number(total_ppn_match.group(1))
            print(f"[✅ PPN by 'Total PPN'] {ppn_val:,.2f}")
            return ppn_val, format_currency(ppn_val)

        for line in relevant_lines:
            text = line.lower()
            if 'ppn' in text and 'ppnbm' not in text:
                match = re.search(r'([\d.,]+)', line)
                if match:
                    ppn = round(clean_number(match.group(1)))
                    print(f"[✅ PPN by keyword] {ppn:,} ← {line}")
                    return ppn, format_currency(ppn)

            if 'harga jual' in text or 'penggantian' in text:
                match = re.search(r'([\d.,]+)', line)
                if match:
                    harga_jual = clean_number(match.group(1))
                    print(f"[🟡 Harga Jual Detected] {harga_jual:,.2f} ← {line}")

            match_all = re.findall(r'([\d.,]{7,})', line)
            for m in match_all:
                val = clean_number(m)
                if val > 1_000_000:
                    angka_terdeteksi.append(val)

        if harga_jual and dpp and harga_jual > dpp:
            ppn = round(harga_jual - dpp)
            print(f"[🟠 Fallback PPN = Harga Jual - DPP] {ppn:,}")
            return ppn, format_currency(ppn)

        if not harga_jual and angka_terdeteksi and dpp > 0:
            kandidat = max(angka_terdeteksi)
            if kandidat > dpp:
                ppn = round(kandidat - dpp)
                print(f"[🔵 PPN dari Max-Angka - DPP] {ppn:,}")
                return ppn, format_currency(ppn)

        print("[⚠️ PPN tidak dapat ditentukan]")
        return 0.0, format_currency(0.0)
    except Exception as e:
        print(f"[ERROR extract_ppn] {e}")
        return 0.0, format_currency(0.0)
    
def preprocess_for_ocr(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 11, 17, 17)  # reduce noise
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    return thresh

def simpan_preview_image(pil_image, halaman_ke):
    """
    Simpan file preview hanya jika belum ada (hindari duplikat)
    """
    pil_image = pil_image.convert("RGB")
    buffer = BytesIO()
    pil_image.save(buffer, format="JPEG", quality=80)
    img_bytes = buffer.getvalue()
    img_hash = hashlib.md5(img_bytes).hexdigest()
    filename = f"{img_hash}_halaman_{halaman_ke}.jpg"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    if not os.path.exists(filepath):
        with open(filepath, 'wb') as f:
            f.write(img_bytes)
        print(f"[📸 PREVIEW DISIMPAN] {filename}")
    else:
        print(f"[📎 PREVIEW SUDAH ADA] {filename}")

    return filename

import json

# Tambahkan ke app.py, ubah function process_file agar support multi-halaman hasil OCR
@app.route('/api/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return jsonify(error="File tidak ditemukan"), 400

    file = request.files['file']
    nama_pt_utama = request.form.get('nama_pt_utama', '').strip()

    if not nama_pt_utama:
        return jsonify(error="Nama PT Utama wajib diisi"), 400

    if not allowed_file(file.filename):
        return jsonify(error="File tidak didukung. Gunakan PDF atau gambar"), 400
    
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        if file.filename.lower().endswith('.pdf'):
            images = convert_from_path(filepath, poppler_path=POPPLER_PATH)
        else:
            with Image.open(filepath) as img:
                images = [img.copy()]  # salinan aman

        hasil_semua_halaman = []

        for i, image in enumerate(images):
            halaman_ke = i + 1
            print(f"\n=== [📝 DEBUG] MEMPROSES HALAMAN {halaman_ke} ===")

            if isinstance(image, np.ndarray):
                img_cv = image
            else:
                img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

            thresh = preprocess_for_ocr(img_cv)

            preview_filename = simpan_preview_image(image, halaman_ke)
            
            raw_text = pytesseract.image_to_string(img_cv, lang='ind', config='--psm 6')
            no_faktur, tanggal_obj = extract_faktur_tanggal(raw_text)
            
            jenis_pajak, blok_rekanan, _ = extract_jenis_pajak(raw_text, nama_pt_utama)
            if not jenis_pajak:
                hasil_semua_halaman.append({"error": f"Hal {halaman_ke}: Nama PT Utama tidak ditemukan."})
                continue

            nama_rekanan, npwp_rekanan = extract_npwp_nama_rekanan(blok_rekanan)
            if not no_faktur or not tanggal_obj:
                hasil_semua_halaman.append({"error": f"Hal {halaman_ke}: Tidak ditemukan tanggal/faktur"})
                continue

            dpp, dpp_str, override_ppn, override_ppn_str = extract_dpp(raw_text)
            ppn, ppn_str = extract_ppn(raw_text, dpp, override_ppn)
            keterangan = extract_keterangan(raw_text)

            hasil_halaman = {
                "klasifikasi": jenis_pajak,
                "data": {
                    "bulan": tanggal_obj.strftime("%B"),
                    "tanggal": tanggal_obj.strftime("%Y-%m-%d"),
                    "keterangan": keterangan,
                    "npwp_lawan_transaksi": npwp_rekanan,
                    "nama_lawan_transaksi": nama_rekanan,
                    "no_faktur": no_faktur,
                    "dpp": dpp,
                    "dpp_str": dpp_str,
                    "ppn": ppn,
                    "ppn_str": ppn_str,
                    "formatted_dpp": dpp_str,
                    "formatted_ppn": ppn_str,
                    "halaman": halaman_ke,
                    "preview_image": preview_filename,
                    "raw_ocr": raw_text
                }
            }

            hasil_semua_halaman.append(hasil_halaman)

            # Simpan debug JSON dan TXT
            debug_dir = os.path.join(UPLOAD_FOLDER, "debug")
            os.makedirs(debug_dir, exist_ok=True)
            json_path = os.path.join(debug_dir, f"{os.path.splitext(file.filename)[0]}_hal_{halaman_ke}.json")
            txt_path = os.path.join(debug_dir, f"debug_page_{halaman_ke}.txt")

            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(hasil_halaman, f, ensure_ascii=False, indent=2)
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(raw_text)

            print(f"[✅ HALAMAN {halaman_ke}] Faktur: {no_faktur} | DPP: {dpp_str} | PPN: {ppn_str}")

            preview_filename = f"preview_{os.path.splitext(file.filename)[0]}_hal_{halaman_ke}.jpg"
            preview_path = os.path.join(UPLOAD_FOLDER, preview_filename)
            cv2.imwrite(preview_path, thresh)

            print("[DEBUG] Hasil halaman ke", halaman_ke)
            print(json.dumps(hasil_semua_halaman[-1], indent=2, ensure_ascii=False))

        return jsonify({
            "success": True,
            "results": hasil_semua_halaman,
            "total_halaman": len(images)
        }), 200

    except Exception as err:
        print(f"[❌ ERROR] {traceback.format_exc()}")
        return jsonify(error=traceback.format_exc()), 500

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

# ==============================================================================
# ENDPOINT BARU: HANYA UNTUK MENYIMPAN DATA
# ==============================================================================
@app.route('/api/save', methods=['POST'])
def save_data():
    data = request.get_json()
    jenis_pajak = data.get('klasifikasi')
    detail_data = data.get('data')

    if not all([jenis_pajak, detail_data]):
        return jsonify(error="Data tidak lengkap"), 400

    model_to_use = PpnMasukan if jenis_pajak == 'PPN_MASUKAN' else PpnKeluaran

    # Cek duplikasi
    existing_record = db.session.execute(db.select(model_to_use).filter_by(no_faktur=detail_data['no_faktur'])).scalar_one_or_none()
    if existing_record:
        return jsonify(message=f"Error: Faktur dengan nomor '{detail_data['no_faktur']}' sudah ada."), 409

    # Simpan data baru
    new_record = model_to_use(
        bulan=detail_data['bulan'],
        tanggal=datetime.strptime(detail_data['tanggal'], '%Y-%m-%d').date(),
        keterangan=detail_data['keterangan'],
        npwp_lawan_transaksi=detail_data['npwp_lawan_transaksi'],
        nama_lawan_transaksi=detail_data['nama_lawan_transaksi'],
        no_faktur=detail_data['no_faktur'],
        dpp=detail_data['dpp'],
        ppn=detail_data['ppn']
    )
    db.session.add(new_record)
    db.session.commit()

    return jsonify(message="Data berhasil disimpan ke database!"), 201

from openpyxl import Workbook

@app.route('/api/export', methods=['GET'])
def export_excel():
    try:

        wb = Workbook()
        ws = wb.active

        masukan = PpnMasukan.query.all()
        keluaran = PpnKeluaran.query.all()
        print("[DEBUG] Masukan:", len(masukan), "| Keluaran:", len(keluaran))

        def serialize_row(row):
            print(f"[DEBUG] Serialize: {row.id} | {row.no_faktur}")
            dpp_val = float(row.dpp)
            ppn_val = float(row.ppn)
            return [
                row.tanggal.strftime('%Y-%m-%d'),
                "PPN MASUKAN" if isinstance(row, PpnMasukan) else "PPN KELUARAN",
                row.keterangan,
                row.npwp_lawan_transaksi,
                row.nama_lawan_transaksi,
                row.no_faktur,
                dpp_val,
                ppn_val,
                dpp_val + ppn_val  # 👈 Tambahan di akhir
            ]

        headers = ['Tanggal', 'Jenis', 'Keterangan', 'NPWP Rekanan', 'Nama Rekanan', 'No Faktur', 'DPP', 'PPN', 'Jumlah (DPP + PPN)']
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_rows = []
        for r in masukan + keluaran:
            try:
                data_rows.append(serialize_row(r))
            except Exception as e:
                print(f"[ERROR] Gagal serialize row ID={r.id}: {e}")

        # Load template
        template_path = os.path.join(os.getcwd(), 'rekap_template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.active

        start_row = 2  # Anggap data mulai dari baris ke-3
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for idx, row_data in enumerate(data_rows, start=start_row):
            for col_index, value in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=col_index, value=value)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                if isinstance(value, (int, float)) and col_index in [7,8,9]:
                    cell.number_format = '#,##0.00'

        # Simpan ke file sementara
        temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp.name)
        print("[DEBUG] Berhasil menyimpan file:", temp.name)
        return send_file(temp.name, as_attachment=True, download_name="rekap_pajak.xlsx")

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/history', methods=['GET'])
def get_history():
    masukan = PpnMasukan.query.all()
    keluaran = PpnKeluaran.query.all()

    def serialize(row, jenis):
        return {
            "id": row.id,
            "jenis": jenis,
            "no_faktur": row.no_faktur,
            "nama_lawan_transaksi": row.nama_lawan_transaksi,
            "tanggal": row.tanggal.strftime('%Y-%m-%d'),
        }

    hasil = [serialize(r, 'masukan') for r in masukan] + [serialize(r, 'keluaran') for r in keluaran]
    return jsonify(hasil), 200

@app.route('/api/delete/<string:jenis>/<int:id>', methods=['DELETE'])
def delete_faktur(jenis, id):
    model = PpnMasukan if jenis.lower() == 'masukan' else PpnKeluaran
    faktur = db.session.get(model, id)
    if not faktur:
        return jsonify(message="Data tidak ditemukan"), 404

    db.session.delete(faktur)
    db.session.commit()
    return jsonify(message="Faktur berhasil dihapus!"), 200

@app.route('/preview/<filename>')
def serve_preview(filename):
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    return send_file(filepath, mimetype='image/jpeg')

# ==============================================================================
# MENJALANKAN APLIKASI (Tidak ada perubahan)
# ==============================================================================
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)