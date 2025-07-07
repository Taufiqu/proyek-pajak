import os
from flask import send_file, current_app
from openpyxl import load_workbook
from io import BytesIO
from models import BuktiSetor
from datetime import datetime

def generate_excel_bukti_setor_export(db):
    try:
        # Ambil data bukti setor dari database
        results = db.session.query(BuktiSetor).order_by(BuktiSetor.tanggal).all()

        # Path absolut ke template
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(BASE_DIR, "..", "..", "templates", "rekap_template_bukti_setor.xlsx")
        template_path = os.path.normpath(template_path)

        # Load template
        wb = load_workbook(template_path)
        ws = wb.active

        # Mulai menulis data dari baris ke-2 (anggap baris 1 adalah header)
        row_num = 2
        for item in results:
            ws.cell(row=row_num, column=1).value = item.tanggal.strftime("%Y-%m-%d")
            ws.cell(row=row_num, column=2).value = item.kode_setor
            ws.cell(row=row_num, column=3).value = float(item.jumlah)
            ws.cell(row=row_num, column=4).value = item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else ""
            row_num += 1

        # Simpan ke buffer memori
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Buat nama file dinamis berdasarkan waktu
        filename = f"bukti_setor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"Gagal export Excel: {e}")
        return {"error": str(e)}, 500